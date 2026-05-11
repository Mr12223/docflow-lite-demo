from __future__ import annotations

import argparse
import csv
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from _bootstrap import ensure_project_root_on_path
from openpyxl import load_workbook

ensure_project_root_on_path()

from docflow.paths import PROJECT_ROOT


DEFAULT_EXCEL_DIR = PROJECT_ROOT / "发票样本" / "invoices" / "excel"
DEFAULT_OUTPUT = PROJECT_ROOT / "发票样本" / "invoice_excel_manifest.csv"
HEADERS = [
    "sample_id",
    "file_name",
    "source_dir",
    "sample_type",
    "expected_invoice",
    "invoice_code",
    "invoice_number",
    "invoice_date",
    "amount",
    "tax",
    "total",
    "buyer_name",
    "seller_name",
    "buyer_tax_id",
    "seller_tax_id",
    "invoice_type",
    "quality_level",
    "notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="generate_invoice_excel_manifest",
        description="根据 Excel 发票模板自动生成字段级评测清单。",
    )
    parser.add_argument(
        "--excel-dir",
        default=str(DEFAULT_EXCEL_DIR),
        help="Excel 发票样本目录，默认 发票样本/invoices/excel。",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="输出清单路径，默认 发票样本/invoice_excel_manifest.csv。",
    )
    return parser.parse_args()


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = _normalize_text(value)
    return text.replace("/", "-").replace(".", "-")


def _normalize_amount(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        amount = Decimal(str(value))
    except InvalidOperation:
        return _normalize_text(value)
    amount = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(amount, "f")


def _cell(ws, ref: str) -> Any:
    return ws[ref].value


def _detect_sample_type(file_name: str) -> str:
    stem = Path(file_name).stem
    if stem.startswith("invoice_electronic_"):
        return "excel_electronic"
    if stem.startswith("invoice_normal_"):
        return "excel_normal"
    if stem.startswith("invoice_receipt_"):
        return "excel_receipt"
    if stem.startswith("invoice_vat_"):
        return "excel_vat"
    return "excel_other"


def _find_invoice_code(ws) -> str:
    for row in ws.iter_rows(min_row=1, max_row=min(6, ws.max_row), values_only=True):
        values = [_normalize_text(value) for value in row]
        for idx, label in enumerate(values):
            if label in {"发票代码", "代码"}:
                for next_idx in range(idx + 1, min(idx + 3, len(values))):
                    candidate = _normalize_text(row[next_idx])
                    if candidate:
                        return candidate
    return ""


def _find_summary_row(ws) -> int | None:
    for row_idx in range(1, ws.max_row + 1):
        first_value = _normalize_text(ws.cell(row_idx, 1).value)
        if first_value.startswith("合计大写"):
            return row_idx

    for row_idx in range(ws.max_row, 0, -1):
        if any(ws.cell(row_idx, col).value not in (None, "") for col in (5, 7, 8)):
            return row_idx
    return None


def extract_row(path: Path, index: int) -> dict[str, str]:
    """Extract ground-truth fields from one Excel invoice template."""

    workbook = load_workbook(path, data_only=True)
    ws = workbook.active
    summary_row = _find_summary_row(ws)
    amount = _normalize_amount(ws.cell(summary_row, 5).value if summary_row else "")
    tax = _normalize_amount(ws.cell(summary_row, 7).value if summary_row else "")
    total = _normalize_amount(ws.cell(summary_row, 8).value if summary_row else "")

    return {
        "sample_id": f"EX{index:03d}",
        "file_name": path.name,
        "source_dir": "invoices/excel",
        "sample_type": _detect_sample_type(path.name),
        "expected_invoice": "true",
        "invoice_code": _find_invoice_code(ws),
        "invoice_number": _normalize_text(_cell(ws, "C2")),
        "invoice_date": _normalize_date(_cell(ws, "G2")),
        "amount": amount,
        "tax": tax,
        "total": total,
        "buyer_name": _normalize_text(_cell(ws, "C3")),
        "seller_name": _normalize_text(_cell(ws, "C4")),
        "buyer_tax_id": _normalize_text(_cell(ws, "G3")).upper(),
        "seller_tax_id": _normalize_text(_cell(ws, "G4")).upper(),
        "invoice_type": _normalize_text(_cell(ws, "A1")),
        "quality_level": "high",
        "notes": "Excel真值自动生成",
    }


def main() -> int:
    args = parse_args()
    excel_dir = Path(args.excel_dir)
    if not excel_dir.is_absolute():
        excel_dir = PROJECT_ROOT / excel_dir
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = PROJECT_ROOT / output_path

    if not excel_dir.exists():
        print(f"未找到 Excel 样本目录：{excel_dir}")
        return 1

    files = sorted(excel_dir.glob("*.xlsx"))
    if not files:
        print(f"目录下没有 Excel 样本：{excel_dir}")
        return 1

    rows = [extract_row(path, index) for index, path in enumerate(files, start=1)]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    type_counter = Counter(row["sample_type"] for row in rows)
    print(f"已生成 Excel 真值清单：{output_path}")
    print(f"- 样本数：{len(rows)}")
    for sample_type, count in sorted(type_counter.items()):
        print(f"- {sample_type}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
